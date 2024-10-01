import pandas as pd
from openpyxl import load_workbook, Workbook
import os

# Define the data structure
data = {
"Main Topic": ["IO", "Operators", "Control Statements", "Arrays", "Strings", "Recursion"],
"Questions": [
"Question 1 for IO", "Question 1 for Operators", "Question 1 for Control Statements",
"Question 1 for Arrays", "Question 1 for Strings", "Question 1 for Recursion"
],
"Constraints": [
"Constraint 1 for IO", "Constraint 1 for Operators", "Constraint 1 for Control Statements",
"Constraint 1 for Arrays", "Constraint 1 for Strings", "Constraint 1 for Recursion"
],
"Input Format": [
"Input format for IO", "Input format for Operators", "Input format for Control Statements",
"Input format for Arrays", "Input format for Strings", "Input format for Recursion"
],
"Output Format": [
"Output format for IO", "Output format for Operators", "Output format for Control Statements",
"Output format for Arrays", "Output format for Strings", "Output format for Recursion"
],
"Test Case 1": ["Test case 1 for IO", "Test case 1 for Operators", "Test case 1 for Control Statements",
"Test case 1 for Arrays", "Test case 1 for Strings", "Test case 1 for Recursion"],
"Test Case 2": ["Test case 2 for IO", "Test case 2 for Operators", "Test case 2 for Control Statements",
"Test case 2 for Arrays", "Test case 2 for Strings", "Test case 2 for Recursion"],
"Test Case 3": ["Test case 3 for IO", "Test case 3 for Operators", "Test case 3 for Control Statements",
"Test case 3 for Arrays", "Test case 3 for Strings", "Test case 3 for Recursion"],
"Test Case 4": ["Test case 4 for IO", "Test case 4 for Operators", "Test case 4 for Control Statements",
"Test case 4 for Arrays", "Test case 4 for Strings", "Test case 4 for Recursion"],
"Test Case 5": ["Test case 5 for IO", "Test case 5 for Operators", "Test case 5 for Control Statements",
"Test case 5 for Arrays", "Test case 5 for Strings", "Test case 5 for Recursion"],
"Test Case 6": ["Test case 6 for IO", "Test case 6 for Operators", "Test case 6 for Control Statements",
"Test case 6 for Arrays", "Test case 6 for Strings", "Test case 6 for Recursion"]
}

# Create a DataFrame
df = pd.DataFrame(data)

# Load or create the Excel file
file_path = "/Users/sivaramangananathan/Desktop/LLMS/react-app/existing_file.xlsx"  # Replace with your actual path
if not os.path.exists(file_path):
# Create a new workbook and save it
    wb = Workbook()
    wb.save(file_path)

# Load the existing Excel file
book = load_workbook(file_path)

# Use the with statement to ensure the file is properly handled
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
df.to_excel(writer, sheet_name="Programming Questions", index=False)

print("Data appended to the existing Excel file successfully!")